import pytesseract
from PIL import Image
from pdf2image import convert_from_path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, numbers
from openpyxl.worksheet.hyperlink import Hyperlink
import os
import shutil
import cv2
import numpy as np
from Alternate_ROI_BLS import Slimming_ROI, clean_text
from datetime import datetime
from Invoice_Reformatter_BLS import invoice_adj
import time
from COMBO_BLS import ocr_text_realignment

# If there are more date formats you are getting, throw them in the date_formats list.

def parse_date(nasty_date):
    date_formats = ["%m/%d/%Y", "%B %d, %Y"]

    for date_format in date_formats:
        try:
            new_date_value = datetime.strptime(nasty_date, date_format)
            converted_date_value = datetime.strptime(str(new_date_value), '%Y-%m-%d %H:%M:%S').strftime('%m/%d/%Y')
            return converted_date_value
        except ValueError:
            pass


def deskew(image):
    image_np = np.array(image)
    gray = cv2.cvtColor(image_np, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    lines = cv2.HoughLines(edges, 1, np.pi / 180, 200)

    if lines is not None:
        angles = []
        for rho, theta in lines[0]:
            if theta < np.pi / 4 or theta > 3 * np.pi / 4:
                angles.append(theta)

        if len(angles) > 0:
            avg_angle = np.mean(angles)
            if not np.isnan(avg_angle):
                avg_angle_deg = np.degrees(avg_angle)
                rotated_image = Image.fromarray(image_np).rotate(-avg_angle_deg, resample=Image.BICUBIC, expand=True)
                return rotated_image

    return image
# To set PATH for Python
pytesseract.pytesseract.tesseract_cmd = r"external_items/Tesseract/tesseract.exe"
# Specify the folder path containing the subfolders containing the PDF files to be scanned

main_folder_path = input("Main invoice path.")
# This is where you put the temporary images that's created during the ocr. Make sure this is empty because the whole folder will be cleared when its complete.
temp_folder = r"temp_images"
# This is the excel file where the ocr results will be sent to. Make sure its blank (the adjusting sheet will refer to it.
excel_file = input("Raw voucher data excel path.")

vendor_info_file_path = input('Enter vendor info path.')
vendor_info_workbook = load_workbook(vendor_info_file_path)
vendor_info_worksheet = vendor_info_workbook['Sheet1']

term_list = vendor_info_worksheet["C"]
id_list = vendor_info_worksheet["B"]
vendor_list = vendor_info_worksheet["A"]

# Define the ROI sets with their corresponding names. ROI stands for range of interest, the area that will be read. You can find this using FindPdfImageCoords code.
roi_sets = {
    'Amazon Business': [
        'Amazoninv_adj',  # ROI 1: Top left (x1, y1), Bottom right (x2, y2)
        (760, 672, 972, 747),  # ROI 2: Top left (x1, y1), Bottom right (x2, y2)
        'Amazondt_adj',
        (1121, 285, 1624, 635),

    ],
    'Red Table': [
        (221, 756, 318, 786), # Invoice I added 150 to the x values
        (585, 687, 826, 723), # Invoice Amount
        (339, 687, 547, 724), # Invoice Date
        (592, 719, 825, 748), # Description

    ],
    'AM Exclusive': [
        (1368, 218, 1490, 244),
        # I removed the parenthesis because its apparently redundant when using the Alternate_ROI method,
        # but put it back if there are issues.
        'amexclusive_amt_adj',
        (1379, 259, 1494, 285),
        (250, 877, 634, 1754),

    ],
    'Eastern Paint': [
        (1357, 148, 1496, 187),
        (1095, 1581, 1651, 1629),
        (1201, 190, 1500, 227),
        (36, 847, 738, 1528),

    ],
    'Tandym': [
        (1036, 319, 1329, 364),
        (1334, 519, 1608, 563),
        (1334, 320, 1608, 362),
        (51, 751, 814, 963),

    ],
    'Rael Fire': [
        (1443, 269, 1614, 302),
        (1310, 394, 1626, 442),
        (1329, 835, 1618, 888),
        (190, 1124, 1063, 1551),

    ],
    'Watch Guard': [
        (1294, 62, 1546, 109),
        'watch_guard_adj',
        (1294, 115, 1546, 161),
        (63, 695, 717, 730),

        # Add more ROI sets as needed
    ],
    'Herff Jones': [
        (845, 229, 1098, 263),
        'herff_amount_adj',
        (1247, 156, 1380, 208),
        (84, 869, 1100, 1350),

        # Add more ROI sets as needed
    ],
    'Mileo Systems': [
        (1452, 227, 1645, 285),
        (1304, 1941, 1645, 2033),
        (1240, 228, 1447, 283),
        (240, 927, 1134, 1934),

    ],
    'IRL Systems': [
        (1420, 223, 1598, 306),
        (1340, 2078, 1622, 2132),
        (1143, 221, 1314, 303),
        (213, 828, 1210, 2071),


    ],
    'Doyle Security Services': [
        (1286, 139, 1375, 180),
        (1193, 599, 1625, 656),
        (1270, 179, 1451, 230),
        (72, 782, 1144, 1757),

    ],
    'Robert Half': [
        (1334, 123, 1494, 158),
        (1442, 1186, 1517, 1228),
        (1338, 92, 1493, 123),
        (128, 694, 626, 766),

    ],
    '1099MI': [
        (459, 330, 704, 359),
        (114, 384, 796, 426),
        (826, 201, 966, 241),
        (1082, 611, 1196, 639),

    ],
    '1099NE': [
        (461, 330, 799, 358),
        (116, 437, 721, 481),
        (826, 268, 939, 306),
        (1082, 611, 1196, 639),

    ],

    # 'Joint': [
    #     (669, 408, 993, 443),
    #     (1132, 1573, 1336, 1615),
    #     (1420, 260, 1665, 303),
    #     (40, 747, 368, 787),
    #
    # ],
    # This was for the October project. Just comment back in if needed, or better yet,
    # use space for non invoice scanning.


}

# If certain invoices need to be reformatted to fit or look a specific way, add it here.
# If it's not here, it'll just pass it normally.
complete = "00000000000000"
# Consider the invoice # from left to right to visualize where to put Xs.
# Example: 'Monarch': "X000000000XXX0" or 'Colonial ES': complete
invoice_formats = {
    'Amazon Business': complete,
}

# Define the names of the ROIs
roi_names = ['Invoice', 'Invoice Amount', 'Invoice Date', 'Description']  # Add more names as needed

# Define the desired data type for each column
data_types = ['text', 'text', 'date', 'text']  # Add more data types as needed

start_time = time.time()
# Create an Excel workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write the ROI headers to the first row of the sheet
header_row = roi_names + ['GL Code']+['AP Type']+['Vendor ID']+['Voucher Date']+['Vendor Name']  # Add 'Folder Name' to the header row
sheet.append(header_row)

# Add a new header for the hyperlink and terms column
sheet.cell(row=1, column=len(header_row) + 1, value='Terms')
sheet.cell(row=1, column=len(header_row) + 2, value='Scanned File')

one_page_invoices = ["Amazon Business", 'AM Exclusive', 'Mileo Systems', 'Tandym', 'Doyle Security Services', 'Red Table', 'Robert Half']

# Format the header row
for cell in sheet[1]:
    cell.number_format = numbers.FORMAT_TEXT  # Apply text formatting to the header cells
    cell.border = Border(bottom=Side(border_style="thin"))

# Iterate through the subfolders
image = []

for subfolder_name in os.listdir(main_folder_path):
    subfolder_path = os.path.join(main_folder_path, subfolder_name)
    if os.path.isdir(subfolder_path):
        folder_name = subfolder_name  # Store the name of the folder
        # Check if the subfolder name matches any ROI set
        if subfolder_name in roi_sets:
            roi_coords = roi_sets[subfolder_name]

            # Iterate through the files in the subfolder
            for file_name in os.listdir(subfolder_path):
                if file_name.endswith((".pdf", ".PDF")):

                    file_path = os.path.join(subfolder_path, file_name)

                    # Convert PDF pages to images
                    images = convert_from_path(file_path, use_pdftocairo=True, output_folder=temp_folder, poppler_path=r"external_items/poppler-24.08.0/Library/bin")

                    # Iterate through the pages and extract text from each ROI
                    image_list = range(len(images))
                    for page_num in image_list:

                        # This if else is to ignore every page other than the first. Assuming that's the only page that
                        # matters, keep this. But if invoices that require multiple pages being scanned show up, you'll
                        # need to add some sort of modifier that only does this if it belongs to certain vendors.
                        print("---------------------------------------------------------------------------------------")
                        print(subfolder_name)

                        if subfolder_name in one_page_invoices and image_list.index(page_num) > 0:
                            pass

                        else:
                            image = images[page_num]
                            roi_texts = []  # Store the extracted text for each ROI in a list
                            for roi, data_type, roi_name in zip(roi_coords, data_types, roi_names):

# ----------------------------------------------------------------------------------------------------------------------

                                if roi == ("Amazoninv_adj"):
                                    replaced_text = Slimming_ROI(roi, "Amazoninv_adj", "Invoice #", image, 1, 268,
                                                                 (1112, 148, 1359, 180), 0)
                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)
                                        except ValueError:
                                            print("There was an error.")
                                            roi_texts.append("")
                                    else:
                                        roi_texts.append(replaced_text)
                            # Make sure to add this whole code block for each elif otherwise amount column will be ignored.
    # ----------------------------------------------------------------------------------------------------------------------
                                elif roi == ("Amazondt_adj"):
                                    replaced_text = Slimming_ROI(roi, "Amazondt_adj", "|", image, 1, 300,
                                                                 (1403, 147, 1618, 183),0)

                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)

                                        except ValueError:
                                            print("There was an error.")
                                            roi_texts.append("")
                                    else:
                                        roi_texts.append(replaced_text)
    # ----------------------------------------------------------------------------------------------------------------------
                                elif roi == ("watch_guard_adj"):
                                    replaced_text = Slimming_ROI(roi, "watch_guard_adj",
                                                                 "Due($)", image, 1, 400, (1,2,3,4), 70)
                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)
                                        except ValueError:
                                            roi_texts.append("")
                                    else:
                                        roi_texts.append(replaced_text)
    # ----------------------------------------------------------------------------------------------------------------------
                                elif roi == ("amexclusive_amt_adj"):
                                    replaced_text = Slimming_ROI(roi, "amexclusive_amt_adj",
                                                                 "Balance Due", image, 1, 200, (1512, 328, 1636, 368), 0)
                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)
                                        except ValueError:
                                            roi_texts.append("")
                                    else:
                                        roi_texts.append(replaced_text)
                    # Add as many of these placeholder things for the special roi until i figure out a prettier way.
 # ---------------------------------------------------------------------------------------------------------------------
                                elif roi == ("herff_amount_adj"):
                                    replaced_text = Slimming_ROI(roi, "herff_amount_adj",
                                                                 "USD", image, 1, 200, (1316, 247, 1438, 306), 0)
                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)
                                        except ValueError:
                                            roi_texts.append("")
                                    else:
                                        roi_texts.append(replaced_text)
                    # Add as many of these placeholder things for the special roi until i figure out a prettier way.
 # ---------------------------------------------------------------------------------------------------------------------
                                elif roi == ("Place_Holder_ROI_Name"):
                                    replaced_text = Slimming_ROI(roi, "Place_Holder_ROI_Name",
                                                                 "Place_Holder_Target_Text", image, 1, 200, (1, 2, 3, 4), 0)
                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)
                                        except ValueError:
                                            roi_texts.append("")
                                    else:
                                        roi_texts.append(replaced_text)
                    # Add as many of these placeholder things for the special roi until i figure out a prettier way.
 # ---------------------------------------------------------------------------------------------------------------------
                                else:
                                    # Crop the image to the specified ROI
                                    roi_image = image.crop(roi)

                                    # Perform OCR on the ROI image and clean the extracted text

                                    # Haters will call this lazy since im explicitly doing this for amazon instead of
                                    # making a list that amazon is in. They'd be right!
                                    if roi == roi_sets["Amazon Business"][3]:
                                        # ocr_dict_text = pytesseract.image_to_data(roi_image,lang="eng",output_type=pytesseract.Output.DICT)
                                        # print(ocr_dict_text)
                                        cleaned_text = ocr_text_realignment(roi_image)
                                        print(cleaned_text)
                                    else:
                                        ocr_text = pytesseract.image_to_string(roi_image)
                                        cleaned_text = clean_text(ocr_text)

                                    # Apply replacement of "$" with "S" to the extracted text
                                    replaced_text = cleaned_text

                                    # Check the data type and apply appropriate formatting
                                    if data_type == 'number':
                                        try:
                                            value = float(replaced_text)
                                            roi_texts.append(value)
                                        except ValueError:
                                            roi_texts.append("")
                                    elif data_type == 'date':
                                        try:
                                            date_value = parse_date(replaced_text)
                                            roi_texts.append(date_value)
                                        except ValueError:
                                            print("There was an error")
                                            roi_texts.append("")
                                    elif roi_name == "Invoice" and data_type == 'text':
                                        adjusted_text = invoice_adj(replaced_text, invoice_formats.get(subfolder_name, ""))
                                        roi_texts.append(adjusted_text)

                                    else:
                                        roi_texts.append(replaced_text)



                            # roi_texts.append(folder_name)
                            # Append folder_name to the end of roi_texts

                            # Write the extracted text for the current page to the next row of the sheet
                            sheet.append(roi_texts)

                            # Add a blank row between pages
                            sheet.append([])

                            # Get the current row number
                            current_row = sheet.max_row

                            # Create the hyperlink to the scanned file and put the file name.
                            # Also add the Vendor ID to the list. Avoiding index/match hurts! My most faithful friend...
                            for vendor in vendor_list:
                                if vendor.value == folder_name:
                                    pos = vendor_list.index(vendor)
                                    vendor_found_id = id_list[pos].value
                                    terms_value = term_list[pos].value
                                    terms_value_cell = sheet.cell(row=current_row, column=len(header_row)+1)
                                    terms_value_cell.value = terms_value
                                    vendor_id_cell = sheet.cell(row=current_row, column=len(header_row)-2)
                                    vendor_id_cell.value = vendor_found_id
                                    print(f"Vendor ID is: {vendor_found_id}")
                                else:
                                    pass

                            folder_cell = sheet.cell(row=current_row, column=len(header_row))
                            folder_cell.value = folder_name
                            file_hyperlink = f'HYPERLINK("{file_path}")'
                            file_cell = sheet.cell(row=current_row, column=len(header_row) + 2)
                            file_cell.value = file_hyperlink
                            file_cell.hyperlink = Hyperlink(ref=f'{file_cell.coordinate}', target=file_path)

                    # Clear the image list to release memory

                    del images[:]
                    for filename in os.listdir(temp_folder):
                        file_path = os.path.join(temp_folder, filename)

                        # Check if it's a file
                        if os.path.isfile(file_path):
                            # Delete files
                            os.remove(file_path)
                        elif os.path.isdir(file_path):
                            # Delete directories and their contents
                            shutil.rmtree(file_path)

# Apply formatting to the cells based on the data types
for i, data_type in enumerate(data_types):
    column = sheet[chr(65 + i)]  # Get the column by its index (A, B, C, etc.)
    for cell in column[1:]:
        if data_type == 'date':
            cell.number_format = "mm/dd/yyyy"
        elif data_type == 'number':
            cell.number_format = numbers.FORMAT_NUMBER_00
        elif data_type == 'text':
            cell.number_format = numbers.FORMAT_TEXT

# Auto-adjust column width to fit the content
for column in sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column[0].column_letter].width = adjusted_width

# Save the workbook to a file

workbook.save(excel_file)

end_time = time.time()
print("Data saved to Excel file:", excel_file)
elapsed_time = end_time - start_time
print("Elapsed time: ", elapsed_time)
os.startfile(excel_file)
# There's not enough memory for more than 90 pages at a time.
